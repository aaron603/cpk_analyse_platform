"""
data_extractor.py
-----------------
Traverses local test station folders, finds test records matching barcodes,
selects the latest successful test for each barcode, copies xlsx/json files
to the output directories, and generates a missing-barcodes Excel report.

Real-world folder layout observed:
  <station_root>/
    [X11|X11_X11|...]/          ← optional grouping sub-folder
      <barcode> or <bc1_bc2>/   ← barcode folder (name *contains* the barcode)
        <YYYYMMDDHHMMSS>/        ← one folder per test attempt
          Test_Result_<ts>_<bc>.xlsx
          <bc>_..._MEASUREMENT_...json
          ...
"""

import json as _json_mod
import os
import re
import shutil
import pandas as pd
from collections import defaultdict
from datetime import datetime


# ---------------------------------------------------------------------------
# Barcode input helpers
# ---------------------------------------------------------------------------

def read_barcodes(excel_path: str) -> list:
    """Read barcode list from 'PrdSN' column of the input Excel file."""
    df = pd.read_excel(excel_path)
    col_map = {c.strip().lower(): c for c in df.columns}
    key = col_map.get('prdsn')
    if key is None:
        raise ValueError(
            f"Input Excel has no 'PrdSN' column. Found: {list(df.columns)}"
        )
    barcodes = df[key].dropna().astype(str).str.strip().tolist()
    return [b for b in barcodes if b]


# ---------------------------------------------------------------------------
# Timestamp parsing
# ---------------------------------------------------------------------------

_TS_PATTERNS = [
    r'(?<!\d)(\d{4})(\d{2})(\d{2})(\d{2})(\d{2})(\d{2})(?!\d)',
    r'(?<!\d)(\d{4})(\d{2})(\d{2})[_\-](\d{2})(\d{2})(\d{2})(?!\d)',
    r'(\d{4})[_\-](\d{2})[_\-](\d{2})[_ ](\d{2})[_\-](\d{2})[_\-](\d{2})',
    r'(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2}):(\d{2})',
]


def _parse_timestamp(name: str) -> datetime:
    for pattern in _TS_PATTERNS:
        m = re.search(pattern, name)
        if m:
            try:
                g = m.groups()
                return datetime(int(g[0]), int(g[1]), int(g[2]),
                                int(g[3]), int(g[4]), int(g[5]))
            except ValueError:
                continue
    return datetime.min


def _is_timestamp_folder(name: str) -> bool:
    return _parse_timestamp(name) != datetime.min


# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------

def _open_excel_safe(path: str):
    try:
        return pd.ExcelFile(path)
    except Exception:
        return None


def is_test_successful(xlsx_path: str) -> bool:
    """Return True iff no 'fail' result in any sheet (case-insensitive)."""
    xl = _open_excel_safe(xlsx_path)
    if xl is None:
        return False
    try:
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            if 'result' not in df.columns:
                continue
            results = df['result'].dropna().astype(str).str.strip().str.lower()
            if (results == 'fail').any():
                return False
        return True
    except Exception:
        return False


def get_earliest_start_time(xlsx_path: str) -> datetime:
    xl = _open_excel_safe(xlsx_path)
    if xl is None:
        return datetime.min
    try:
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            if 'start_time' not in df.columns:
                continue
            times = pd.to_datetime(df['start_time'], errors='coerce').dropna()
            if not times.empty:
                return times.min().to_pydatetime()
    except Exception:
        pass
    return datetime.min


def _find_xlsx_for_barcode(folder: str, barcode: str):
    candidates = []
    try:
        for name in os.listdir(folder):
            if not name.lower().endswith('.xlsx'):
                continue
            full = os.path.join(folder, name)
            if os.path.isfile(full):
                if barcode in name:
                    return full
                candidates.append(full)
    except OSError:
        pass
    return candidates[0] if candidates else None


def _find_measurement_json(folder: str, barcode: str):
    candidates = []
    try:
        for name in os.listdir(folder):
            if not name.lower().endswith('.json'):
                continue
            full = os.path.join(folder, name)
            if os.path.isfile(full):
                if 'DutInfo' in name:
                    continue
                if barcode in name:
                    return full
                if 'MEASUREMENT' in name.upper():
                    candidates.append(full)
    except OSError:
        pass
    return candidates[0] if candidates else None


# ---------------------------------------------------------------------------
# Core discovery
# ---------------------------------------------------------------------------

def find_test_records(barcode: str, station_root: str) -> list:
    """
    Walk station_root and return all test records for barcode.

    Supports any layout where, at some depth:
      - A folder whose name contains the barcode (the "barcode folder")
      - Has timestamp-named (YYYYMMDDHHMMSS) subdirectories directly beneath it
      - Each timestamp subfolder contains the xlsx/json test files

    Standard structure (confirmed):
      <root>/TestResult/<product>/<station>/[variable middle levels]/<barcode>/<timestamp>/files

    The number of levels between <station> and <barcode> may vary per engineer.
    Legacy structure (also supported):
      <root>/[<grouping>/]<barcode_or_combined>/<timestamp>/files

    Rules:
      - Folders named "debug" (case-insensitive) are always skipped.
      - Non-data helper dirs (file_bk, RU1_Log_*, TM1_Log, ...) are skipped
        once we are inside a timestamp folder.
      - Once a barcode-match folder is found, we handle its timestamp
        subdirs directly and stop recursing deeper into it.

    Each record: (ts_folder_path, xlsx_path, json_path, is_success, test_time)
    xlsx_path may be None for json-only records.
    """
    records = []
    norm_root = os.path.normpath(station_root)

    # Folder names to prune (partial match prefix or exact name, lower-case)
    _SKIP_EXACT = {'debug', 'file_bk', 'env_comp', 'tm1_log'}
    _SKIP_PREFIX = ('ru1_log_', 'ru2_log_', 'tm1_log', 'tm2_log',
                    'gain flatness', 'tx aclr', 'peak power',
                    'rx sensitivity', 'efficency')

    def _should_skip_dir(name: str) -> bool:
        nl = name.lower()
        return nl in _SKIP_EXACT or any(nl.startswith(p) for p in _SKIP_PREFIX)

    for root, dirs, files in os.walk(norm_root):
        norm_root_r = os.path.normpath(root)
        rel = os.path.relpath(norm_root_r, norm_root)
        depth = 0 if rel == '.' else rel.count(os.sep) + 1

        # Hard depth limit — allow up to 10 to accommodate variable intermediate levels
        if depth > 10:
            dirs.clear()
            continue

        folder_name = os.path.basename(norm_root_r)

        # Always prune skip-list dirs before recursing
        dirs[:] = [d for d in dirs if not _should_skip_dir(d)]

        # If we are inside a timestamp folder, stop recursing (nothing useful deeper)
        if _is_timestamp_folder(folder_name):
            dirs.clear()
            continue

        # If the current folder name contains the barcode, process its
        # timestamp subdirs and stop descending into this folder.
        if barcode in folder_name:
            for ts_dir in list(dirs):
                if not _is_timestamp_folder(ts_dir):
                    continue
                ts_path = os.path.join(norm_root_r, ts_dir)
                xlsx = _find_xlsx_for_barcode(ts_path, barcode)
                json_f = _find_measurement_json(ts_path, barcode)

                if xlsx is not None:
                    success = is_test_successful(xlsx)
                    test_time = get_earliest_start_time(xlsx)
                elif json_f is not None:
                    success = _json_result_pass(json_f)
                    test_time = _json_start_time(json_f)
                else:
                    continue

                if test_time == datetime.min:
                    test_time = _parse_timestamp(ts_dir)

                records.append((ts_path, xlsx, json_f, success, test_time))

            # Do not descend further into this barcode folder
            dirs.clear()

    return records


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def _json_result_pass(json_path: str) -> bool:
    for enc in ('utf-8', 'utf-8-sig', 'gbk', 'latin-1'):
        try:
            with open(json_path, encoding=enc) as f:
                data = _json_mod.load(f)
            result = data.get('DutInfo', {}).get('Result', 'Fail')
            return str(result).strip().lower() == 'pass'
        except UnicodeDecodeError:
            continue
        except Exception:
            return False
    return False


def _json_start_time(json_path: str) -> datetime:
    for enc in ('utf-8', 'utf-8-sig', 'gbk', 'latin-1'):
        try:
            with open(json_path, encoding=enc) as f:
                data = _json_mod.load(f)
            ts_str = data.get('DutInfo', {}).get('StartTime', '')
            if ts_str:
                return datetime.strptime(ts_str[:19], '%Y-%m-%d %H:%M:%S')
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    return datetime.min


# ---------------------------------------------------------------------------
# Public extraction entry point
# ---------------------------------------------------------------------------

def run_extraction(
    barcodes: list,
    station_configs: list,
    output_base_dir: str,
    log_cb=None,
    progress_cb=None,
    stop_event=None,
    mode: str = 'latest_pass',
) -> dict:
    """
    Run full extraction for all barcodes across all station types.

    mode:
      'latest_pass' — (default) only the newest successful record per barcode
      'all_pass'    — all successful records per barcode (multiple xlsx copied)
      'all'         — all records regardless of pass/fail

    Multiple configs with the same 'type' are merged.
    Result dict per barcode includes extra fields for the missing report.
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    # Group folders by station type
    type_to_folders: dict = defaultdict(list)
    for cfg in station_configs:
        stype = cfg.get('type', '').strip()
        sfolder = cfg.get('folder', '').strip()
        if stype and sfolder:
            type_to_folders[stype].append(sfolder)

    summary = {}
    n_types = len(type_to_folders)
    total = len(barcodes) * n_types
    done = 0

    for stype, folders in type_to_folders.items():
        if stop_event and stop_event.is_set():
            _log(f"[INFO] 用户中止，跳过剩余工站")
            break
        valid_folders = [f for f in folders if os.path.isdir(f)]
        invalid_folders = [f for f in folders if not os.path.isdir(f)]

        if invalid_folders:
            for f in invalid_folders:
                _log(f"  [WARN] 工站 [{stype}] 文件夹不存在，已跳过: {f}")

        if not valid_folders:
            _log(f"[SKIP] 工站 [{stype}]: 无有效文件夹，跳过该工站")
            continue

        xlsx_out = os.path.join(output_base_dir, stype, 'xlsx')
        json_out = os.path.join(output_base_dir, stype, 'json')
        # Clear existing output dirs to avoid accumulating files from prior runs
        for _d in (xlsx_out, json_out):
            if os.path.exists(_d):
                shutil.rmtree(_d)
            os.makedirs(_d)

        _log(f"\n{'='*60}")
        _log(f"[INFO] 开始处理工站类型: [{stype}]")
        _log(f"[INFO] 共配置 {len(valid_folders)} 个文件夹:")
        for f in valid_folders:
            _log(f"       → {f}")
        _log(f"[INFO] 待查询条码: {len(barcodes)} 个")
        _log(f"[INFO] 输出目录: xlsx → {xlsx_out}")
        _log(f"{'='*60}")

        results = []
        not_found_n, no_pass_n, no_xlsx_n, ok_n, err_n = 0, 0, 0, 0, 0

        for bc in barcodes:
            all_records = []
            read_errors = []

            for folder in valid_folders:
                try:
                    recs = find_test_records(bc, folder)
                    all_records.extend(recs)
                except Exception as exc:
                    read_errors.append(f"{folder}: {exc}")

            if read_errors:
                for err in read_errors:
                    _log(f"  [ERROR] {bc} 读取文件夹时出错 — {err}")
                err_n += 1

            # Build summary fields
            total_recs = len(all_records)
            pass_recs = sum(1 for _, _, _, s, _ in all_records if s)
            latest_any = max(
                (t for _, _, _, _, t in all_records), default=datetime.min
            )
            latest_any_str = (
                latest_any.strftime('%Y-%m-%d %H:%M:%S')
                if latest_any != datetime.min else ''
            )

            if not all_records:
                not_found_n += 1
                _log(f"  [WARN] {bc} — 未找到任何测试记录")
                r = {
                    'status': 'not_found', 'barcode': bc,
                    'message': '未找到任何测试记录',
                    'xlsx': None, 'json': None,
                    'total_records': 0, 'pass_records': 0,
                    'latest_any_time': '', 'found_in': '', 'note': '',
                }

            else:
                # ── Select candidate records based on mode ────────────
                if mode == 'all':
                    # All records that have xlsx, regardless of pass/fail
                    candidates = [
                        (p, x, j, s, t) for p, x, j, s, t in all_records
                        if x is not None
                    ]
                elif mode == 'fail_only':
                    # Only failed records with xlsx
                    candidates = [
                        (p, x, j, s, t) for p, x, j, s, t in all_records
                        if not s and x is not None
                    ]
                else:
                    # latest_pass / all_pass: only successful xlsx records
                    candidates = [
                        (p, x, j, s, t) for p, x, j, s, t in all_records
                        if s and x is not None
                    ]
                    all_pass_any = any(s for _, _, _, s, _ in all_records)

                if mode in ('latest_pass', 'all_pass') and not all_pass_any:
                    no_pass_n += 1
                    _log(
                        f"  [WARN] {bc} — 找到 {total_recs} 条记录，"
                        f"通过 {pass_recs} 条，失败 {total_recs - pass_recs} 条，"
                        f"无通过记录，最近测试: {latest_any_str}"
                    )
                    r = {
                        'status': 'no_pass', 'barcode': bc,
                        'message': f'找到{total_recs}条记录，均未全部通过',
                        'xlsx': None, 'json': None,
                        'total_records': total_recs, 'pass_records': pass_recs,
                        'latest_any_time': latest_any_str, 'found_in': '', 'note': '',
                    }

                elif not candidates:
                    # Has records but no usable xlsx
                    no_xlsx_n += 1
                    reason = (
                        '无失败记录（均通过）' if mode == 'fail_only'
                        else 'json-only记录，无xlsx' if mode != 'all'
                        else 'json-only，无xlsx'
                    )
                    _log(
                        f"  [WARN] {bc} — 找到 {total_recs} 条记录，"
                        f"但全为json-only（无xlsx），无法进行 CPK 分析"
                    )
                    r = {
                        'status': 'no_xlsx', 'barcode': bc,
                        'message': '记录存在但无xlsx，无法CPK分析',
                        'xlsx': None, 'json': None,
                        'total_records': total_recs, 'pass_records': pass_recs,
                        'latest_any_time': latest_any_str, 'found_in': '',
                        'note': reason,
                    }

                else:
                    # Copy one or all candidate records
                    candidates.sort(key=lambda x: x[4], reverse=True)  # newest first

                    to_copy = (
                        candidates[:1]          # latest_pass: single record
                        if mode == 'latest_pass'
                        else candidates         # all_pass / all: every record
                    )

                    last_dest_xlsx = None
                    last_dest_json = None
                    copy_errors = []
                    found_in = ''

                    for ts_folder, src_xlsx, src_json, rec_pass, test_time in to_copy:
                        dest_xlsx = os.path.join(xlsx_out, os.path.basename(src_xlsx))
                        try:
                            shutil.copy2(src_xlsx, dest_xlsx)
                            last_dest_xlsx = dest_xlsx
                        except Exception as exc:
                            copy_errors.append(f'xlsx复制失败: {exc}')
                            _log(f"  [ERROR] {bc} xlsx 复制失败: {exc}")
                            continue

                        if src_json:
                            dest_json = os.path.join(
                                json_out, os.path.basename(src_json)
                            )
                            try:
                                shutil.copy2(src_json, dest_json)
                                last_dest_json = dest_json
                            except Exception as exc:
                                _log(f"  [WARN] {bc} json 复制失败: {exc}")

                        if not found_in:
                            for folder in valid_folders:
                                if os.path.normpath(folder) in \
                                        os.path.normpath(ts_folder):
                                    found_in = folder
                                    break

                    ok_n += 1
                    time_str = candidates[0][4].strftime('%Y-%m-%d %H:%M:%S')
                    count_note = (
                        f"（共{total_recs}条记录，复制{len(to_copy)}条）"
                        if len(to_copy) > 1 else
                        f"（共{total_recs}条记录，取最新）" if total_recs > 1 else ''
                    )
                    pass_flag = candidates[0][3]
                    status_tag = '[OK]  ' if pass_flag else '[OK-F]'  # F=fail record
                    _log(
                        f"  {status_tag} {bc}  测试时间:{time_str}"
                        f"{'' if last_dest_json else '（无json）'}{count_note}"
                    )
                    r = {
                        'status': 'success', 'barcode': bc, 'message': 'OK',
                        'xlsx': last_dest_xlsx, 'json': last_dest_json,
                        'total_records': total_recs, 'pass_records': pass_recs,
                        'latest_any_time': time_str, 'found_in': found_in,
                        'note': '; '.join(copy_errors),
                    }

            results.append(r)
            done += 1
            if progress_cb:
                progress_cb(done, total, bc)

            if stop_event and stop_event.is_set():
                _log(f"  [INFO] 用户中止，已处理 {len(results)}/{len(barcodes)} 个条码")
                break

        missing_n = not_found_n + no_pass_n + no_xlsx_n
        _log(f"\n[汇总] 工站 [{stype}]  总条码: {len(barcodes)}")
        _log(f"       成功提取: {ok_n}  |  无通过记录: {no_pass_n}"
             f"  |  无xlsx: {no_xlsx_n}  |  未找到: {not_found_n}"
             f"  |  读取错误: {err_n}")
        _log(f"       缺失（需关注）: {missing_n} 个条码")

        summary[stype] = {
            'xlsx_dir': xlsx_out,
            'json_dir': json_out,
            'results': results,
        }

    return summary


# ---------------------------------------------------------------------------
# Auto-discover barcodes from station folders (used when no input Excel)
# ---------------------------------------------------------------------------

def discover_barcodes(station_folders: list) -> list:
    """
    Walk all configured station folders and return unique barcode strings.

    A "barcode folder" is any folder that directly contains at least one
    timestamp-named subdirectory (YYYYMMDDHHMMSS format). Dual-barcode
    folder names like BC1_BC2 are split into individual barcodes when
    both parts are alphanumeric strings of 5+ characters.

    Skip rules follow the same list as find_test_records().
    """
    _SKIP_EXACT = {'debug', 'file_bk', 'env_comp', 'tm1_log', 'testresult'}
    _SKIP_PREFIX = ('ru1_log_', 'ru2_log_', 'tm1_log', 'tm2_log',
                    'gain flatness', 'tx aclr', 'peak power',
                    'rx sensitivity', 'efficency')

    def _should_skip(name: str) -> bool:
        nl = name.lower()
        return nl in _SKIP_EXACT or any(nl.startswith(p) for p in _SKIP_PREFIX)

    seen: set = set()
    result: list = []

    for root_folder in station_folders:
        if not os.path.isdir(root_folder):
            continue

        for dirpath, dirs, _files in os.walk(root_folder):
            dirs[:] = [d for d in dirs if not _should_skip(d)]
            folder_name = os.path.basename(dirpath)

            # A barcode folder must directly contain at least one timestamp dir
            has_ts_child = any(_is_timestamp_folder(d) for d in dirs)
            if not has_ts_child:
                continue

            # Stop descending into this barcode folder
            dirs.clear()

            # Try to split dual-barcode names (e.g. BC1_BC2)
            parts = folder_name.split('_')
            bc_parts = [p for p in parts if len(p) >= 5 and p.isalnum()]

            if len(bc_parts) >= 2:
                # Dual-barcode folder — add each individually
                for p in bc_parts:
                    if p not in seen:
                        seen.add(p)
                        result.append(p)
            else:
                # Single barcode or short name — add the whole folder name
                if folder_name and folder_name not in seen:
                    seen.add(folder_name)
                    result.append(folder_name)

    return result


# ---------------------------------------------------------------------------
# Missing barcodes Excel report
# ---------------------------------------------------------------------------

_STATUS_LABEL = {
    'not_found': '未找到测试记录',
    'no_pass':   '无通过测试记录',
    'no_xlsx':   '有通过记录但无xlsx',
    'error':     '读取异常',
}


def generate_missing_report(summary: dict, output_path: str, log_cb=None) -> str:
    """
    Generate an Excel report listing all non-success barcodes per station type.

    Each station type gets one sheet.  A summary row is appended at the bottom.
    Returns the absolute path of the written file.
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # Colours
    HDR_FILL   = PatternFill('solid', fgColor='1A237E')
    HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
    WARN_FILL  = PatternFill('solid', fgColor='FFF3E0')
    ERR_FILL   = PatternFill('solid', fgColor='FFEBEE')
    SUM_FILL   = PatternFill('solid', fgColor='E8EAF6')
    SUM_FONT   = Font(bold=True, size=10)
    THIN       = Side(style='thin', color='BBBBBB')
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER     = Alignment(horizontal='center', vertical='center')
    LEFT       = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    COLUMNS = [
        ('PrdSN',       18),
        ('状态',         18),
        ('总测试次数',    10),
        ('通过次数',      10),
        ('最新测试时间',  20),
        ('备注',         40),
    ]

    total_missing_all = 0

    for stype, info in summary.items():
        results   = info['results']
        total_bc  = len(results)
        ok_n      = sum(1 for r in results if r['status'] == 'success')
        missing   = [r for r in results if r['status'] != 'success']
        total_missing_all += len(missing)

        ws = wb.create_sheet(title=stype[:31])   # sheet name max 31 chars

        # ── Header row ──────────────────────────────────────────────────
        for col_idx, (col_name, col_w) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_w

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

        # ── Data rows ────────────────────────────────────────────────────
        for row_idx, r in enumerate(missing, 2):
            status_label = _STATUS_LABEL.get(r['status'], r['status'])
            row_fill = ERR_FILL if r['status'] == 'not_found' else WARN_FILL

            values = [
                r['barcode'],
                status_label,
                r.get('total_records', ''),
                r.get('pass_records', ''),
                r.get('latest_any_time', ''),
                r.get('note', '') or r.get('message', ''),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = row_fill
                cell.border    = BORDER
                cell.alignment = CENTER if col_idx != len(values) else LEFT
                cell.font      = Font(size=10)

        # ── Summary row ──────────────────────────────────────────────────
        sum_row = len(missing) + 2
        ws.cell(row=sum_row, column=1, value='汇总').font = SUM_FONT
        ws.cell(row=sum_row, column=1).fill = SUM_FILL
        ws.cell(row=sum_row, column=1).alignment = CENTER
        summary_text = (
            f"总条码: {total_bc}  |  成功提取: {ok_n}"
            f"  |  缺失/异常: {len(missing)}"
        )
        cell = ws.cell(row=sum_row, column=2, value=summary_text)
        cell.font = SUM_FONT
        cell.fill = SUM_FILL
        cell.alignment = LEFT
        ws.merge_cells(
            start_row=sum_row, start_column=2,
            end_row=sum_row, end_column=len(COLUMNS)
        )
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=sum_row, column=col_idx).border = BORDER

        _log(f"  [缺失报表] 工站 [{stype}]: 总{total_bc}条  成功{ok_n}"
             f"  缺失/异常{len(missing)}条")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    _log(f"\n[INFO] 缺失条码报表已保存: {output_path}"
         f"  (共 {total_missing_all} 个缺失/异常条码)")
    return output_path


# ---------------------------------------------------------------------------
# Duplicate barcode report (all_pass mode)
# ---------------------------------------------------------------------------

def generate_duplicate_report(summary: dict, output_path: str, log_cb=None) -> str:
    """
    Generate an Excel report listing barcodes that have multiple pass records
    in an all_pass extraction.  Each station type gets one sheet.

    Columns: PrdSN | 工站类型 | 重复测试次数 | 各次测试时间
    Only barcodes with count > 1 are listed.
    Returns the path of the written file.
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill('solid', fgColor='1A237E')
    HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
    DUP_FILL = PatternFill('solid', fgColor='FFF8E1')
    SUM_FILL = PatternFill('solid', fgColor='E8EAF6')
    SUM_FONT = Font(bold=True, size=10)
    THIN      = Side(style='thin', color='BBBBBB')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=False)
    LEFT_WRAP = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    COLUMNS = [
        ('PrdSN',       24),
        ('工站类型',     10),
        ('重复测试次数', 12),
        ('各次测试时间', 52),
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_dup_barcodes = 0

    for stype, info in summary.items():
        results = info.get('results', [])

        # Group test times by barcode
        bc_times: dict = {}
        for r in results:
            bc = r['barcode']
            t  = r.get('latest_any_time', '')
            bc_times.setdefault(bc, []).append(t)

        duplicates = {bc: ts for bc, ts in bc_times.items() if len(ts) > 1}
        if not duplicates:
            _log(f'  [重复报表] 工站 [{stype}]: 无重复条码')
            continue

        total_dup_barcodes += len(duplicates)
        ws = wb.create_sheet(title=stype[:31])

        # Header
        for col_idx, (col_name, col_w) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_w
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

        # Data rows — sorted by test count descending
        sorted_dups = sorted(duplicates.items(), key=lambda x: len(x[1]), reverse=True)
        for row_idx, (bc, times) in enumerate(sorted_dups, 2):
            times_sorted = sorted(t for t in times if t)
            times_str = '\n'.join(times_sorted)
            row_h = max(15, 15 * len(times_sorted))

            for col_idx, val in enumerate(
                [bc, stype, len(times_sorted), times_str], 1
            ):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = DUP_FILL
                cell.border    = BORDER
                cell.alignment = LEFT_WRAP if col_idx == 4 else CENTER
                cell.font      = Font(size=10)
            ws.row_dimensions[row_idx].height = row_h

        # Summary row
        sum_row = len(sorted_dups) + 2
        total_rec = sum(len(ts) for ts in duplicates.values())
        xlsx_n = sum(1 for r in results if r.get('xlsx'))
        json_n = sum(1 for r in results if r.get('json'))
        ws.cell(row=sum_row, column=1, value='汇总').font = SUM_FONT
        ws.cell(row=sum_row, column=1).fill      = SUM_FILL
        ws.cell(row=sum_row, column=1).alignment = CENTER
        summary_text = (
            f'工站 [{stype}]  总提取记录: {len(results)}'
            f'  (xlsx: {xlsx_n}, json: {json_n})'
            f'  |  重复条码数: {len(duplicates)}'
            f'  |  重复记录总数: {total_rec}'
        )
        cell = ws.cell(row=sum_row, column=2, value=summary_text)
        cell.font      = SUM_FONT
        cell.fill      = SUM_FILL
        cell.alignment = LEFT_WRAP
        ws.merge_cells(
            start_row=sum_row, start_column=2,
            end_row=sum_row,   end_column=len(COLUMNS)
        )
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=sum_row, column=col_idx).border = BORDER

        _log(f'  [重复报表] 工站 [{stype}]: 重复条码 {len(duplicates)} 个，'
             f'共 {total_rec} 条重复记录  |  xlsx文件: {xlsx_n}  json文件: {json_n}')

    # ── Sheet: xlsx_json不一致 ──────────────────────────────────────────────
    MISMATCH_FILL_XLSX = PatternFill('solid', fgColor='E3F2FD')  # xlsx-only: blue tint
    MISMATCH_FILL_JSON = PatternFill('solid', fgColor='F3E5F5')  # json-only: purple tint

    mismatch_rows = []
    for stype, info in summary.items():
        results = info.get('results', [])
        for r in results:
            has_xlsx = bool(r.get('xlsx'))
            has_json = bool(r.get('json'))
            if has_xlsx and not has_json:
                mismatch_rows.append((r['barcode'], stype, '有xlsx，无json', ''))
            elif has_json and not has_xlsx:
                mismatch_rows.append((r['barcode'], stype, '有json，无xlsx', ''))

    if mismatch_rows:
        ws_m = wb.create_sheet(title='xlsx_json不一致')
        COLS_M = [('PrdSN', 24), ('工站类型', 12), ('不一致类型', 20), ('备注', 30)]
        for ci, (cname, cw) in enumerate(COLS_M, 1):
            c = ws_m.cell(row=1, column=ci, value=cname)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
            ws_m.column_dimensions[get_column_letter(ci)].width = cw
        ws_m.row_dimensions[1].height = 20
        ws_m.freeze_panes = 'A2'

        for ri, (bc, stype, mis_type, note) in enumerate(mismatch_rows, 2):
            row_fill = (MISMATCH_FILL_XLSX if '有xlsx' in mis_type
                        else MISMATCH_FILL_JSON)
            for ci, val in enumerate([bc, stype, mis_type, note], 1):
                c = ws_m.cell(row=ri, column=ci, value=val)
                c.fill = row_fill; c.border = BORDER
                c.font = Font(size=10)
                c.alignment = LEFT_WRAP if ci == 1 else CENTER

        sum_r = len(mismatch_rows) + 2
        xlsx_only_cnt = sum(1 for _, _, t, _ in mismatch_rows if '有xlsx' in t)
        json_only_cnt = len(mismatch_rows) - xlsx_only_cnt
        ws_m.cell(row=sum_r, column=1, value='汇总').font = SUM_FONT
        ws_m.cell(row=sum_r, column=1).fill = SUM_FILL
        ws_m.cell(row=sum_r, column=1).alignment = CENTER
        mis_summary = (f'共 {len(mismatch_rows)} 个条码不一致'
                       f'  |  有xlsx无json: {xlsx_only_cnt}'
                       f'  |  有json无xlsx: {json_only_cnt}')
        c = ws_m.cell(row=sum_r, column=2, value=mis_summary)
        c.font = SUM_FONT; c.fill = SUM_FILL; c.alignment = LEFT_WRAP
        ws_m.merge_cells(start_row=sum_r, start_column=2,
                         end_row=sum_r, end_column=len(COLS_M))
        for ci in range(1, len(COLS_M) + 1):
            ws_m.cell(row=sum_r, column=ci).border = BORDER

        _log(f'  [xlsx_json不一致] 共 {len(mismatch_rows)} 个条码不一致'
             f'（有xlsx无json: {xlsx_only_cnt}，有json无xlsx: {json_only_cnt}）')
    else:
        _log('  [xlsx_json对比] 所有条码xlsx与json文件均一一对应，无不一致')

    # ── Sheets: 不完整测试条码 / 完整测试条码 ──────────────────────────────
    try:
        from core.cpk_calculator import analyze_xlsx_completeness as _check_complete
    except ImportError:
        try:
            from cpk_calculator import analyze_xlsx_completeness as _check_complete
        except ImportError:
            _check_complete = None

    if _check_complete:
        COMPLETE_FILL   = PatternFill('solid', fgColor='E8F5E9')  # green
        INCOMPL_FILL    = PatternFill('solid', fgColor='FFF8E1')  # yellow
        MISSING_FILL    = PatternFill('solid', fgColor='FFEBEE')  # red

        all_incomplete_rows = []   # (barcode, stype, sheet, pname, present_n, ref_n, fname)
        all_complete_rows   = []   # (barcode, stype, n_items, fname)

        for stype, info in summary.items():
            xlsx_dir = info.get('xlsx_dir', '')
            if not xlsx_dir or not os.path.isdir(xlsx_dir):
                continue
            comp = _check_complete(xlsx_dir, log_cb=_log)
            ref_n = len(comp['reference_set'])

            for item in comp['complete']:
                all_complete_rows.append(
                    (item['barcode'], stype, item['n_items'], item['filename'])
                )
            for item in comp['incomplete']:
                for (sheet_name, pname) in item['missing']:
                    all_incomplete_rows.append((
                        item['barcode'], stype,
                        sheet_name, pname,
                        item['present_n'], ref_n,
                        item['filename'],
                    ))

        # ── Sheet: 不完整测试条码 ────────────────────────────────────────────
        if all_incomplete_rows:
            ws_inc = wb.create_sheet(title='不完整测试条码')
            COLS_INC = [
                ('PrdSN',         24), ('工站',         10),
                ('测试大项',       28), ('缺失测试子项',  30),
                ('已有子项数',     12), ('参考完整子项数', 14),
                ('文件名',         36),
            ]
            for ci, (cname, cw) in enumerate(COLS_INC, 1):
                c = ws_inc.cell(row=1, column=ci, value=cname)
                c.fill = HDR_FILL; c.font = HDR_FONT
                c.alignment = CENTER; c.border = BORDER
                ws_inc.column_dimensions[get_column_letter(ci)].width = cw
            ws_inc.row_dimensions[1].height = 20
            ws_inc.freeze_panes = 'A2'

            # Sort: most missing items first, then by barcode
            from collections import Counter as _Ctr
            bc_miss_cnt = _Ctr(r[0] for r in all_incomplete_rows)
            all_incomplete_rows.sort(
                key=lambda r: (-bc_miss_cnt[r[0]], r[0], r[2], r[3])
            )

            for ri, (bc, st, sh, pn, pres, ref, fn) in enumerate(
                    all_incomplete_rows, 2):
                for ci, val in enumerate(
                        [bc, st, sh, pn, pres, ref, fn], 1):
                    c = ws_inc.cell(row=ri, column=ci, value=val)
                    c.fill = MISSING_FILL if ci == 4 else INCOMPL_FILL
                    c.border = BORDER
                    c.font = Font(size=10)
                    c.alignment = (LEFT_WRAP if ci in (1, 3, 4, 7) else CENTER)

            inc_bc_cnt = len(set(r[0] for r in all_incomplete_rows))
            sum_r = len(all_incomplete_rows) + 2
            ws_inc.cell(row=sum_r, column=1, value='汇总').font = SUM_FONT
            ws_inc.cell(row=sum_r, column=1).fill = SUM_FILL
            ws_inc.cell(row=sum_r, column=1).alignment = CENTER
            inc_txt = (f'共 {inc_bc_cnt} 个条码/测试文件存在缺失子项'
                       f'  |  共 {len(all_incomplete_rows)} 条缺失记录')
            c = ws_inc.cell(row=sum_r, column=2, value=inc_txt)
            c.font = SUM_FONT; c.fill = SUM_FILL; c.alignment = LEFT_WRAP
            ws_inc.merge_cells(start_row=sum_r, start_column=2,
                               end_row=sum_r, end_column=len(COLS_INC))
            for ci in range(1, len(COLS_INC) + 1):
                ws_inc.cell(row=sum_r, column=ci).border = BORDER

            _log(f'  [不完整条码] 共 {inc_bc_cnt} 个条码/文件存在测试子项缺失'
                 f'，{len(all_incomplete_rows)} 条缺失记录')
        else:
            _log('  [完整性检查] 所有条码测试子项均完整，无缺失')

        # ── Sheet: 完整测试条码 ──────────────────────────────────────────────
        if all_complete_rows:
            ws_comp = wb.create_sheet(title='完整测试条码')
            COLS_COMP = [
                ('PrdSN',    24), ('工站',     10),
                ('子项总数', 12), ('文件名',   36),
            ]
            for ci, (cname, cw) in enumerate(COLS_COMP, 1):
                c = ws_comp.cell(row=1, column=ci, value=cname)
                c.fill = HDR_FILL; c.font = HDR_FONT
                c.alignment = CENTER; c.border = BORDER
                ws_comp.column_dimensions[get_column_letter(ci)].width = cw
            ws_comp.row_dimensions[1].height = 20
            ws_comp.freeze_panes = 'A2'

            all_complete_rows.sort(key=lambda r: (r[1], r[0]))  # sort by stype, barcode
            for ri, (bc, st, n_items, fn) in enumerate(all_complete_rows, 2):
                for ci, val in enumerate([bc, st, n_items, fn], 1):
                    c = ws_comp.cell(row=ri, column=ci, value=val)
                    c.fill = COMPLETE_FILL; c.border = BORDER
                    c.font = Font(size=10)
                    c.alignment = LEFT_WRAP if ci in (1, 4) else CENTER

            sum_r = len(all_complete_rows) + 2
            ws_comp.cell(row=sum_r, column=1, value='汇总').font = SUM_FONT
            ws_comp.cell(row=sum_r, column=1).fill = SUM_FILL
            ws_comp.cell(row=sum_r, column=1).alignment = CENTER
            c = ws_comp.cell(row=sum_r, column=2,
                             value=f'共 {len(all_complete_rows)} 个条码/文件测试子项完整')
            c.font = SUM_FONT; c.fill = SUM_FILL; c.alignment = LEFT_WRAP
            ws_comp.merge_cells(start_row=sum_r, start_column=2,
                                end_row=sum_r, end_column=len(COLS_COMP))
            for ci in range(1, len(COLS_COMP) + 1):
                ws_comp.cell(row=sum_r, column=ci).border = BORDER

            _log(f'  [完整条码] 共 {len(all_complete_rows)} 个条码/文件测试子项完整')

    if not wb.sheetnames:
        ws = wb.create_sheet(title='无重复条码')
        ws.cell(row=1, column=1, value='所有条码均为单次测试，无重复数据')

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    _log(f'\n[INFO] 重复条码报表已保存: {output_path}'
         f'  (共 {total_dup_barcodes} 个重复条码)')
    return output_path


# ---------------------------------------------------------------------------
# Direct walk extraction (all_pass mode — no barcode list needed)
# ---------------------------------------------------------------------------

def _walk_all_pass_in_folder(station_root: str) -> list:
    """
    Walk station_root and return ALL pass test records for every barcode found.
    No barcode list required — discovers all barcode folders in the tree.

    Returns list of (ts_folder_path, xlsx_path, json_path, test_time, barcode).
    xlsx_path or json_path may be None.
    """
    records = []
    norm_root = os.path.normpath(station_root)

    _SKIP_EXACT = {'debug', 'file_bk', 'env_comp', 'tm1_log'}
    _SKIP_PREFIX = ('ru1_log_', 'ru2_log_', 'tm1_log', 'tm2_log',
                    'gain flatness', 'tx aclr', 'peak power',
                    'rx sensitivity', 'efficency')

    def _should_skip(name: str) -> bool:
        nl = name.lower()
        return nl in _SKIP_EXACT or any(nl.startswith(p) for p in _SKIP_PREFIX)

    for root, dirs, _files in os.walk(norm_root):
        norm_r = os.path.normpath(root)
        rel = os.path.relpath(norm_r, norm_root)
        depth = 0 if rel == '.' else rel.count(os.sep) + 1

        if depth > 10:
            dirs.clear()
            continue

        folder_name = os.path.basename(norm_r)
        dirs[:] = [d for d in dirs if not _should_skip(d)]

        # Once inside a timestamp folder, stop descending
        if _is_timestamp_folder(folder_name):
            dirs.clear()
            continue

        # A barcode folder has at least one timestamp subdir
        ts_dirs = [d for d in dirs if _is_timestamp_folder(d)]
        if not ts_dirs:
            continue

        barcode = folder_name  # folder name IS the barcode (single or dual)

        for ts_dir in ts_dirs:
            ts_path = os.path.join(norm_r, ts_dir)
            xlsx = _find_xlsx_for_barcode(ts_path, barcode)
            json_f = _find_measurement_json(ts_path, barcode)

            if xlsx is not None:
                success = is_test_successful(xlsx)
                test_time = get_earliest_start_time(xlsx)
            elif json_f is not None:
                success = _json_result_pass(json_f)
                test_time = _json_start_time(json_f)
            else:
                continue

            if not success:
                continue  # only collect pass records

            if test_time == datetime.min:
                test_time = _parse_timestamp(ts_dir)

            records.append((ts_path, xlsx, json_f, test_time, barcode))

        # Don't recurse into this barcode folder
        dirs.clear()

    return records


def run_extraction_all_pass(
    station_configs: list,
    output_base_dir: str,
    log_cb=None,
    progress_cb=None,
    stop_event=None,
) -> dict:
    """
    Directly walk all configured station folders and extract every pass record.
    No barcode list is needed.  Returns the same summary structure as run_extraction.
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    type_to_folders: dict = defaultdict(list)
    for cfg in station_configs:
        stype = cfg.get('type', '').strip()
        sfolder = cfg.get('folder', '').strip()
        if stype and sfolder:
            type_to_folders[stype].append(sfolder)

    summary = {}

    for stype, folders in type_to_folders.items():
        if stop_event and stop_event.is_set():
            _log('[INFO] 用户中止，跳过剩余工站')
            break

        valid_folders = [f for f in folders if os.path.isdir(f)]
        invalid_folders = [f for f in folders if not os.path.isdir(f)]

        for f in invalid_folders:
            _log(f'  [WARN] 工站 [{stype}] 文件夹不存在，已跳过: {f}')

        if not valid_folders:
            _log(f'[SKIP] 工站 [{stype}]: 无有效文件夹，跳过该工站')
            continue

        xlsx_out = os.path.join(output_base_dir, stype, 'xlsx')
        json_out = os.path.join(output_base_dir, stype, 'json')
        for _d in (xlsx_out, json_out):
            if os.path.exists(_d):
                shutil.rmtree(_d)
            os.makedirs(_d)

        _log(f"\n{'='*60}")
        _log(f'[INFO] 开始遍历工站 [{stype}]，提取全部 pass 记录')
        _log(f'[INFO] 共配置 {len(valid_folders)} 个文件夹:')
        for f in valid_folders:
            _log(f'       → {f}')
        _log(f"{'='*60}")

        # ── Collect all pass records from every valid folder ─────────
        all_recs = []
        for folder in valid_folders:
            if stop_event and stop_event.is_set():
                break
            try:
                recs = _walk_all_pass_in_folder(folder)
                _log(f'  [INFO] {folder}')
                _log(f'         发现 {len(recs)} 条 pass 记录')
                all_recs.extend(recs)
            except Exception as exc:
                _log(f'  [ERROR] 遍历文件夹出错 {folder}: {exc}')

        _log(f'  [INFO] 工站 [{stype}] 合计 {len(all_recs)} 条 pass 记录，开始复制...')

        # ── Copy xlsx and json to output dirs ─────────────────────────
        ok_n = 0
        err_n = 0
        results = []
        total = len(all_recs)

        for i, (ts_path, src_xlsx, src_json, test_time, barcode) in enumerate(all_recs):
            if stop_event and stop_event.is_set():
                _log(f'  [INFO] 用户中止，已处理 {i}/{total} 条记录')
                break

            copied_xlsx = None
            copied_json = None

            if src_xlsx:
                dest = os.path.join(xlsx_out, os.path.basename(src_xlsx))
                try:
                    shutil.copy2(src_xlsx, dest)
                    copied_xlsx = dest
                except Exception as exc:
                    _log(f'  [ERROR] xlsx 复制失败 {os.path.basename(src_xlsx)}: {exc}')
                    err_n += 1
                    continue

            if src_json:
                dest = os.path.join(json_out, os.path.basename(src_json))
                try:
                    shutil.copy2(src_json, dest)
                    copied_json = dest
                except Exception as exc:
                    _log(f'  [WARN] json 复制失败: {exc}')

            ok_n += 1
            time_str = (test_time.strftime('%Y-%m-%d %H:%M:%S')
                        if test_time != datetime.min else '')
            results.append({
                'status': 'success',
                'barcode': barcode,
                'message': 'OK',
                'xlsx': copied_xlsx,
                'json': copied_json,
                'total_records': 1,
                'pass_records': 1,
                'latest_any_time': time_str,
                'found_in': ts_path,
                'note': '',
            })

            if progress_cb:
                progress_cb(i + 1, total, barcode)

        xlsx_backed_n = sum(1 for r in results if r.get('xlsx'))
        json_backed_n = sum(1 for r in results if r.get('json'))
        both_n        = sum(1 for r in results if r.get('xlsx') and r.get('json'))
        xlsx_only_n   = sum(1 for r in results if r.get('xlsx') and not r.get('json'))
        json_only_n   = sum(1 for r in results if r.get('json') and not r.get('xlsx'))

        _log(f'\n[汇总] 工站 [{stype}]  成功复制: {ok_n} 条  |  复制错误: {err_n} 条')
        _log(f'       xlsx文件: {xlsx_backed_n}  |  json文件: {json_backed_n}'
             f'  |  均有: {both_n}  |  仅xlsx: {xlsx_only_n}  |  仅json: {json_only_n}')
        if xlsx_only_n or json_only_n:
            _log(f'  [注意] 有 {xlsx_only_n + json_only_n} 个条码xlsx与json文件不对应，'
                 f'详见重复条码报表"xlsx_json不一致"Sheet')

        summary[stype] = {
            'xlsx_dir': xlsx_out,
            'json_dir': json_out,
            'results': results,
        }

    return summary


# ---------------------------------------------------------------------------
# Folder-direct mode helpers (Scenario B: multi-level traverse + extract)
# ---------------------------------------------------------------------------

def check_has_direct_files(folder: str, file_type: str = 'xlsx') -> bool:
    """
    Return True if folder directly contains at least one target file
    (xlsx or json depending on file_type).  Used to detect Scenario A vs B.
    """
    ext = '.xlsx' if file_type == 'xlsx' else '.json'
    try:
        for name in os.listdir(folder):
            if name.lower().endswith(ext) and os.path.isfile(os.path.join(folder, name)):
                return True
    except OSError:
        pass
    return False


def _walk_all_records_in_folder(station_root: str, log_cb=None) -> list:
    """
    Walk station_root and return ALL test records (pass AND fail) for every
    barcode found.  No barcode list required — discovers all barcode folders.

    Returns list of (ts_folder_path, xlsx_path, json_path, test_time, barcode, is_pass).
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    records = []
    norm_root = os.path.normpath(station_root)

    if not os.path.isdir(norm_root):
        _log(f'  [ERROR] 工站文件夹不存在或不是目录: {norm_root}')
        return records

    _log(f'  [WALK] 开始扫描: {norm_root}')

    _SKIP_EXACT = {'debug', 'file_bk', 'env_comp', 'tm1_log'}
    _SKIP_PREFIX = ('ru1_log_', 'ru2_log_', 'tm1_log', 'tm2_log',
                    'gain flatness', 'tx aclr', 'peak power',
                    'rx sensitivity', 'efficency')

    def _should_skip(name: str) -> bool:
        nl = name.lower()
        return nl in _SKIP_EXACT or any(nl.startswith(p) for p in _SKIP_PREFIX)

    for root, dirs, files in os.walk(norm_root):
        norm_r = os.path.normpath(root)
        rel = os.path.relpath(norm_r, norm_root)
        depth = 0 if rel == '.' else rel.count(os.sep) + 1

        if depth > 10:
            _log(f'  [WALK] 深度超限(>10)，停止递归: {norm_r}')
            dirs.clear()
            continue

        folder_name = os.path.basename(norm_r)
        before_skip = list(dirs)
        dirs[:] = [d for d in dirs if not _should_skip(d)]
        skipped = set(before_skip) - set(dirs)
        if skipped:
            _log(f'  [WALK] depth={depth} [{folder_name}] 已跳过子目录: {skipped}')

        if _is_timestamp_folder(folder_name):
            # We are inside a timestamp folder — don't recurse further
            dirs.clear()
            continue

        ts_dirs = [d for d in dirs if _is_timestamp_folder(d)]
        if not ts_dirs:
            _log(f'  [WALK] depth={depth} [{folder_name}] '
                 f'无时间戳子目录，继续向下（共{len(dirs)}个子目录: '
                 f'{dirs[:5]}{"..." if len(dirs)>5 else ""}）')
            continue

        # Current folder is a barcode folder — its children are timestamp dirs
        barcode = folder_name
        _log(f'  [WALK] depth={depth} 发现条码文件夹: [{barcode}]，'
             f'含 {len(ts_dirs)} 个时间戳目录')

        found_in_barcode = 0
        for ts_dir in ts_dirs:
            ts_path = os.path.join(norm_r, ts_dir)
            # Use first part of barcode (before '_') for file matching
            # to handle dual-barcode folder names like WV123_WV456
            bc_key = barcode.split('_')[0] if '_' in barcode else barcode
            xlsx = _find_xlsx_for_barcode(ts_path, bc_key)
            if xlsx is None:
                xlsx = _find_xlsx_for_barcode(ts_path, barcode)
            json_f = _find_measurement_json(ts_path, bc_key)
            if json_f is None:
                json_f = _find_measurement_json(ts_path, barcode)

            if xlsx is not None:
                is_pass = is_test_successful(xlsx)
                test_time = get_earliest_start_time(xlsx)
            elif json_f is not None:
                is_pass = _json_result_pass(json_f)
                test_time = _json_start_time(json_f)
            else:
                _log(f'  [WALK]   [{ts_dir}] 未找到xlsx或json文件，跳过'
                     f'（目录内容: {os.listdir(ts_path)[:8] if os.path.isdir(ts_path) else "N/A"}）')
                continue

            if test_time == datetime.min:
                test_time = _parse_timestamp(ts_dir)

            records.append((ts_path, xlsx, json_f, test_time, barcode, is_pass))
            found_in_barcode += 1

        _log(f'  [WALK]   条码 [{barcode}] 共采集 {found_in_barcode} 条记录')
        dirs.clear()  # Do not recurse into timestamp subdirs

    _log(f'  [WALK] 扫描完成，共 {len(records)} 条记录: {norm_root}')
    return records


def _read_fail_items_from_xlsx(xlsx_path: str) -> list:
    """
    Read all failed test points from an xlsx file.
    Returns list of dicts: {sheet, point_name, data, limit_low, limit_high, deviation}.
    """
    xl = _open_excel_safe(xlsx_path)
    if xl is None:
        return []
    fail_items = []
    try:
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            if 'result' not in df.columns:
                continue
            mask = (df['result'].dropna().astype(str).str.strip().str.lower() == 'fail')
            mask = mask.reindex(df.index, fill_value=False)
            for _, row in df[mask].iterrows():
                point_name = (str(row.get('point_name', '')).strip()
                              if 'point_name' in df.columns else '')
                data_val   = row.get('data', '')
                limit_low  = row.get('limit_low', '')
                limit_high = row.get('limit_high', '')
                deviation  = ''
                try:
                    val = float(data_val)
                    lo  = (float(limit_low)
                           if str(limit_low).strip() not in ('', 'nan', 'None') else None)
                    hi  = (float(limit_high)
                           if str(limit_high).strip() not in ('', 'nan', 'None') else None)
                    if lo is not None and val < lo:
                        deviation = f'{val - lo:.4g}'
                    elif hi is not None and val > hi:
                        deviation = f'{val - hi:.4g}'
                except (ValueError, TypeError):
                    pass
                fail_items.append({
                    'sheet':      sheet,
                    'point_name': point_name,
                    'data':       data_val,
                    'limit_low':  limit_low,
                    'limit_high': limit_high,
                    'deviation':  deviation,
                })
    except Exception:
        pass
    return fail_items


def run_extraction_traverse(
    station_configs: list,
    output_base_dir: str,
    file_type: str = 'xlsx',
    log_cb=None,
    progress_cb=None,
    stop_event=None,
    barcodes: list = None,
) -> tuple:
    """
    folder_direct Scenario B: traverse all station folders, collect ALL records
    (pass + fail), copy to output dirs, and build fail analysis data.

    Returns (extraction_summary, fail_data) where:
      extraction_summary  — same format as run_extraction (used for CPK analysis)
      fail_data           — {stype: {barcode_stats, fail_barcodes,
                                     never_pass_barcodes, all_fail_items}}
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    type_to_folders: dict = defaultdict(list)
    for cfg in station_configs:
        stype   = cfg.get('type', '').strip()
        sfolder = cfg.get('folder', '').strip()
        if stype and sfolder:
            type_to_folders[stype].append(sfolder)

    extraction_summary = {}
    fail_data          = {}

    for stype, folders in type_to_folders.items():
        if stop_event and stop_event.is_set():
            _log('[INFO] 用户中止，跳过剩余工站')
            break

        valid_folders   = [f for f in folders if os.path.isdir(f)]
        invalid_folders = [f for f in folders if not os.path.isdir(f)]

        for f in invalid_folders:
            _log(f'  [WARN] 工站 [{stype}] 文件夹不存在，已跳过: {f}')

        if not valid_folders:
            _log(f'[SKIP] 工站 [{stype}]: 无有效文件夹，跳过该工站')
            continue

        xlsx_out = os.path.join(output_base_dir, stype, 'xlsx')
        json_out = os.path.join(output_base_dir, stype, 'json')
        for _d in (xlsx_out, json_out):
            if os.path.exists(_d):
                shutil.rmtree(_d)
            os.makedirs(_d)

        _log(f"\n{'='*60}")
        _log(f'[INFO] 开始遍历工站 [{stype}]，提取全部记录（pass + fail）')
        _log(f'[INFO] 共配置 {len(valid_folders)} 个文件夹:')
        for f in valid_folders:
            _log(f'       → {f}')
        _log(f"{'='*60}")

        all_recs = []
        for folder in valid_folders:
            if stop_event and stop_event.is_set():
                break
            try:
                recs   = _walk_all_records_in_folder(folder, log_cb=_log)
                pass_n = sum(1 for r in recs if r[5])
                fail_n = len(recs) - pass_n
                _log(f'  [INFO] {folder}')
                _log(f'         发现 {len(recs)} 条记录（pass: {pass_n}, fail: {fail_n}）')
                all_recs.extend(recs)
            except Exception as exc:
                _log(f'  [ERROR] 遍历文件夹出错 {folder}: {exc}')

        # ── Barcode filter ────────────────────────────────────────────
        if barcodes:
            barcodes_set = set(barcodes)
            before = len(all_recs)
            # Handle dual-barcode folders (e.g. WV123_WV456): match if any
            # barcode from the list is a substring of the folder name.
            all_recs = [
                r for r in all_recs
                if r[4] in barcodes_set or any(bc in r[4] for bc in barcodes_set)
            ]
            _log(f'  [INFO] 条码过滤: {before} → {len(all_recs)} 条记录'
                 f'（保留 {len(set(r[4] for r in all_recs))} 个条码）')

        pass_total = sum(1 for r in all_recs if r[5])
        fail_total = len(all_recs) - pass_total
        _log(f'  [INFO] 工站 [{stype}] 合计 {len(all_recs)} 条记录'
             f'（pass: {pass_total}, fail: {fail_total}），开始处理...')

        ok_n           = 0
        err_n          = 0
        results        = []
        barcode_stats  = {}   # barcode → {pass_count, fail_count, times, fail_items}
        all_fail_items = []   # (barcode, time_str, sheet, point_name, data, lsl, usl, dev)

        total = len(all_recs)
        for i, (ts_path, src_xlsx, src_json, test_time, barcode, is_pass) in enumerate(all_recs):
            if stop_event and stop_event.is_set():
                _log(f'  [INFO] 用户中止，已处理 {i}/{total} 条记录')
                break

            time_str = (test_time.strftime('%Y-%m-%d %H:%M:%S')
                        if test_time != datetime.min else '')

            if barcode not in barcode_stats:
                barcode_stats[barcode] = {
                    'pass_count': 0, 'fail_count': 0,
                    'times': [], 'fail_items': [],
                }
            stats = barcode_stats[barcode]
            if time_str:
                stats['times'].append(time_str)

            if is_pass:
                stats['pass_count'] += 1
            else:
                stats['fail_count'] += 1
                if src_xlsx:
                    items = _read_fail_items_from_xlsx(src_xlsx)
                    stats['fail_items'].extend(items)
                    for item in items:
                        all_fail_items.append((
                            barcode, time_str,
                            item['sheet'], item['point_name'],
                            item['data'], item['limit_low'],
                            item['limit_high'], item['deviation'],
                        ))

            # Copy files to output dirs
            copied_xlsx = None
            copied_json = None
            if src_xlsx:
                dest = os.path.join(xlsx_out, os.path.basename(src_xlsx))
                try:
                    shutil.copy2(src_xlsx, dest)
                    copied_xlsx = dest
                except Exception as exc:
                    _log(f'  [ERROR] xlsx 复制失败 {os.path.basename(src_xlsx)}: {exc}')
                    err_n += 1

            if src_json:
                dest = os.path.join(json_out, os.path.basename(src_json))
                try:
                    shutil.copy2(src_json, dest)
                    copied_json = dest
                except Exception as exc:
                    _log(f'  [WARN] json 复制失败: {exc}')

            if copied_xlsx is not None:
                ok_n += 1
                results.append({
                    'status':          'success',
                    'barcode':         barcode,
                    'message':         'OK',
                    'xlsx':            copied_xlsx,
                    'json':            copied_json,
                    'total_records':   1,
                    'pass_records':    1 if is_pass else 0,
                    'latest_any_time': time_str,
                    'found_in':        ts_path,
                    'note':            '',
                })

            if progress_cb:
                progress_cb(i + 1, total, barcode)

        fail_barcodes       = {bc: st for bc, st in barcode_stats.items()
                               if st['fail_count'] > 0}
        never_pass_barcodes = [bc for bc, st in barcode_stats.items()
                               if st['pass_count'] == 0]

        _log(f'\n[汇总] 工站 [{stype}]  总记录: {len(all_recs)}'
             f'  复制成功: {ok_n}  |  错误: {err_n}')
        _log(f'       条码总数: {len(barcode_stats)}'
             f'  |  有失败记录: {len(fail_barcodes)}'
             f'  |  从未通过: {len(never_pass_barcodes)}')

        extraction_summary[stype] = {
            'xlsx_dir': xlsx_out,
            'json_dir': json_out,
            'results':  results,
        }
        fail_data[stype] = {
            'barcode_stats':       barcode_stats,
            'fail_barcodes':       fail_barcodes,
            'never_pass_barcodes': never_pass_barcodes,
            'all_fail_items':      all_fail_items,
        }

    return extraction_summary, fail_data


def generate_folder_direct_excel(
    fail_data: dict,
    output_path: str,
    log_cb=None,
) -> str:
    """
    Generate a 3-sheet Excel for folder_direct Scenario B:
      Sheet 1: 失败条码       — barcodes that had at least one fail
      Sheet 2: 失败测试项     — every individual failed test point row
      Sheet 3: 从未成功条码  — barcodes that never passed
    """
    def _log(msg):
        if log_cb:
            log_cb(msg)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill('solid', fgColor='1A237E')
    HDR_FONT  = Font(color='FFFFFF', bold=True, size=10)
    FAIL_FILL = PatternFill('solid', fgColor='FFEBEE')
    ITEM_FILL = PatternFill('solid', fgColor='FCE4EC')
    NEVR_FILL = PatternFill('solid', fgColor='FFF3E0')
    SUM_FILL  = PatternFill('solid', fgColor='E8EAF6')
    SUM_FONT  = Font(bold=True, size=10)
    THIN      = Side(style='thin', color='BBBBBB')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal='center', vertical='center')
    LEFT      = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _hdr_row(ws, cols):
        for ci, (name, width) in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

    def _sum_row(ws, row_i, cols, text):
        ws.cell(row=row_i, column=1, value='汇总').fill = SUM_FILL
        ws.cell(row=row_i, column=1).font = SUM_FONT
        ws.cell(row=row_i, column=1).alignment = CENTER
        ws.cell(row=row_i, column=2, value=text).fill = SUM_FILL
        ws.cell(row=row_i, column=2).font = SUM_FONT
        ws.cell(row=row_i, column=2).alignment = LEFT
        ws.merge_cells(start_row=row_i, start_column=2,
                       end_row=row_i,   end_column=len(cols))
        for ci in range(1, len(cols) + 1):
            ws.cell(row=row_i, column=ci).border = BORDER

    # ── Sheet 1: 失败条码 ──────────────────────────────────────────────────
    ws1 = wb.create_sheet('失败条码')
    COLS1 = [('条码', 24), ('工站', 12), ('失败次数', 10),
             ('通过次数', 10), ('失败测试项数', 14), ('最近测试时间', 20)]
    _hdr_row(ws1, COLS1)
    row1 = 2
    total_fail_bc = 0
    for stype, sdata in fail_data.items():
        for bc, st in sorted(sdata['fail_barcodes'].items()):
            latest = max(st['times']) if st['times'] else ''
            for ci, v in enumerate(
                [bc, stype, st['fail_count'], st['pass_count'],
                 len(st['fail_items']), latest], 1
            ):
                c = ws1.cell(row=row1, column=ci, value=v)
                c.fill = FAIL_FILL; c.border = BORDER
                c.font = Font(size=10)
                c.alignment = LEFT if ci == 1 else CENTER
            row1 += 1
            total_fail_bc += 1
    _sum_row(ws1, row1, COLS1, f'共 {total_fail_bc} 个失败条码')

    # ── Sheet 2: 失败测试项 ────────────────────────────────────────────────
    ws2 = wb.create_sheet('失败测试项')
    COLS2 = [('测试大项', 28), ('测试子项', 28), ('条码', 20),
             ('测量值', 12), ('下限', 10), ('上限', 10), ('偏差', 10), ('测试时间', 20)]
    _hdr_row(ws2, COLS2)
    row2 = 2
    total_items = 0
    for stype, sdata in fail_data.items():
        for rec in sdata['all_fail_items']:
            bc, time_str, sheet, point, data, lsl, usl, dev = rec
            for ci, v in enumerate(
                [sheet, point, bc, data, lsl, usl, dev, time_str], 1
            ):
                c = ws2.cell(row=row2, column=ci, value=v)
                c.fill = ITEM_FILL; c.border = BORDER
                c.font = Font(size=10)
                c.alignment = LEFT if ci in (1, 2, 3) else CENTER
            row2 += 1
            total_items += 1
    _sum_row(ws2, row2, COLS2, f'共 {total_items} 条失败测试记录')

    # ── Sheet 3: 从未成功条码 ──────────────────────────────────────────────
    ws3 = wb.create_sheet('从未成功条码')
    COLS3 = [('条码', 24), ('工站', 12), ('总测试次数', 12), ('最近测试时间', 20)]
    _hdr_row(ws3, COLS3)
    row3 = 2
    total_never = 0
    for stype, sdata in fail_data.items():
        for bc in sorted(sdata['never_pass_barcodes']):
            st = sdata['barcode_stats'].get(bc, {})
            latest = max(st.get('times', [])) if st.get('times') else ''
            total_tests = st.get('fail_count', 0)
            for ci, v in enumerate([bc, stype, total_tests, latest], 1):
                c = ws3.cell(row=row3, column=ci, value=v)
                c.fill = NEVR_FILL; c.border = BORDER
                c.font = Font(size=10)
                c.alignment = LEFT if ci == 1 else CENTER
            row3 += 1
            total_never += 1
    _sum_row(ws3, row3, COLS3, f'共 {total_never} 个从未通过条码')

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    _log(f'\n[INFO] 失败分析Excel已保存: {output_path}')
    _log(f'       Sheet1 失败条码: {total_fail_bc}'
         f'  |  Sheet2 失败测试项: {total_items}'
         f'  |  Sheet3 从未成功: {total_never}')
    return output_path